"""
Create Excel feedback for a single student grading result.

This script takes grading results from RamiAutoGrader and creates a
professionally formatted Excel file for one student.
"""

import os
import sys
from typing import Dict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter


COLUMNS = [
    {"header": "Student Name", "width": 25, "key": "student_name"},
    {"header": "GitHub URL", "width": 50, "key": "github_url"},
    {"header": "Claimed Grade", "width": 15, "key": "claimed_grade"},
    {"header": "Actual Grade", "width": 15, "key": "actual_grade"},
    {"header": "Status", "width": 12, "key": "status"},
    {"header": "Summary", "width": 80, "key": "summary"},
    {"header": "Notes", "width": 20, "key": "notes"}
]


def add_header_row(ws, columns):
    """Add formatted header row to worksheet."""
    header_font = Font(bold=True, size=11, name='Calibri')
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2",
                               fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for idx, col in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=idx)
        cell.value = col["header"]
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(idx)].width = col["width"]


def add_data_row(ws, student_data, columns):
    """Add single student data row to worksheet."""
    cell_font = Font(size=11, name='Calibri')
    cell_alignment = Alignment(horizontal="left", vertical="top",
                                wrap_text=False)
    summary_alignment = Alignment(horizontal="left", vertical="top",
                                   wrap_text=True)
    hyperlink_font = Font(color="0563C1", underline="single", size=11)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    row_idx = 2
    for col_idx, col in enumerate(columns, start=1):
        cell = ws.cell(row=row_idx, column=col_idx)
        value = student_data.get(col["key"], "")

        if col["key"] == "github_url" and value:
            cell.hyperlink = value
            cell.value = "Link"
            cell.font = hyperlink_font
        else:
            cell.value = value
            cell.font = cell_font

        if col["key"] == "summary":
            cell.alignment = summary_alignment
        else:
            cell.alignment = cell_alignment

        cell.border = thin_border


def create_excel_feedback(student_data: Dict, output_dir: str = "results/"):
    """
    Create formatted Excel workbook for single student.

    Args:
        student_data: Dictionary with student grading information
        output_dir: Output directory path

    Returns:
        Path to created Excel file
    """
    os.makedirs(output_dir, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Grading Feedback"

    add_header_row(ws, COLUMNS)
    add_data_row(ws, student_data, COLUMNS)

    # Create filename from student name
    student_name = student_data.get("student_name", "Student")
    safe_name = "".join(c for c in student_name if c.isalnum() or c in (' ', '-', '_'))
    safe_name = safe_name.replace(' ', '_')
    filename = f"{safe_name}_Feedback.xlsx"
    filepath = os.path.join(output_dir, filename)

    wb.save(filepath)
    return filepath


if __name__ == "__main__":
    import argparse
    import json

    parser = argparse.ArgumentParser(description="Create Excel feedback for student grading")
    parser.add_argument("--json", help="Path to JSON file with student data")
    parser.add_argument("--student-name", help="Student name")
    parser.add_argument("--github-url", help="GitHub repository URL")
    parser.add_argument("--claimed-grade", help="Claimed grade (e.g., 80/100)")
    parser.add_argument("--actual-grade", help="Actual grade (e.g., 87/100)")
    parser.add_argument("--status", help="PASSED or FAILED")
    parser.add_argument("--summary", help="Feedback summary")
    parser.add_argument("--notes", default="", help="Additional notes")
    parser.add_argument("--output-dir", default="results/", help="Output directory")

    args = parser.parse_args()

    if args.json:
        with open(args.json, 'r') as f:
            student_data = json.load(f)
    else:
        student_data = {
            "student_name": args.student_name or "Student",
            "github_url": args.github_url or "",
            "claimed_grade": args.claimed_grade or "Not specified",
            "actual_grade": args.actual_grade or "0/100",
            "status": args.status or "UNKNOWN",
            "summary": args.summary or "",
            "notes": args.notes
        }

    filepath = create_excel_feedback(student_data, args.output_dir)
    print(f"[OK] Excel file created: {filepath}")
