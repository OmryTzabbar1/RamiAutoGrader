"""
Unit tests for populate-excel skill.

Tests cover:
- Excel creation with valid data
- Single student, multiple students
- Header formatting
- Column widths
- Hyperlink creation
- Text wrapping
- Borders
- Error handling (empty data, permission errors)
"""

import os
import pytest
import tempfile
import shutil
from openpyxl import load_workbook
import sys

sys.path.insert(0, os.path.join(os.path.dirname(__file__), '../../skills/populate-excel/scripts'))

from create_excel import populate_excel, add_header_row, add_data_rows, COLUMNS


class TestPopulateExcel:
    """Test Excel creation workflow."""

    @pytest.fixture
    def temp_dir(self):
        """Create temporary directory for test outputs."""
        temp_dir = tempfile.mkdtemp()
        yield temp_dir
        shutil.rmtree(temp_dir)

    @pytest.fixture
    def sample_student_data(self):
        """Sample student data for testing."""
        return [
            {
                "student_id": "12345678",
                "student_name": "John Doe",
                "partner_name": "Jane Smith",
                "assignment_name": "Design Patterns",
                "github_url": "https://github.com/johndoe/project",
                "grade": "77/100",
                "summary": "Score: 77/100. Strong docs, needs improvement in code quality.",
                "notes": ""
            },
            {
                "student_id": "87654321",
                "student_name": "Alice Johnson",
                "partner_name": "Bob Williams",
                "assignment_name": "Design Patterns",
                "github_url": "https://github.com/alicejohnson/patterns",
                "grade": "92/100",
                "summary": "Score: 92/100. Exceptional code quality and perfect test coverage.",
                "notes": ""
            }
        ]

    def test_create_excel_valid_data(self, temp_dir, sample_student_data):
        """Test creating Excel with valid student data."""
        filepath = populate_excel(
            student_data=sample_student_data,
            assignment_name="Design Patterns",
            output_dir=temp_dir
        )

        assert os.path.exists(filepath)
        assert filepath.endswith("FinalFeedback_DesignPatterns.xlsx")

        wb = load_workbook(filepath)
        ws = wb.active

        assert ws.title == "FinalFeedback"
        assert ws.max_row == 3  # Header + 2 students
        assert ws.max_column == 8  # 8 columns

    def test_create_excel_single_student(self, temp_dir):
        """Test creating Excel with single student."""
        student_data = [
            {
                "student_id": "12345678",
                "student_name": "John Doe",
                "partner_name": "Jane Smith",
                "assignment_name": "Design Patterns",
                "github_url": "https://github.com/johndoe/project",
                "grade": "85/100",
                "summary": "Score: 85/100. Good work overall.",
                "notes": ""
            }
        ]

        filepath = populate_excel(student_data, "Test Assignment", temp_dir)

        assert os.path.exists(filepath)

        wb = load_workbook(filepath)
        ws = wb.active

        assert ws.max_row == 2  # Header + 1 student
        assert ws['A2'].value == "12345678"
        assert ws['B2'].value == "John Doe"

    def test_create_excel_many_students(self, temp_dir):
        """Test creating Excel with 30 students."""
        student_data = [
            {
                "student_id": f"1234567{i}",
                "student_name": f"Student {i}",
                "partner_name": f"Partner {i}",
                "assignment_name": "Test",
                "github_url": f"https://github.com/student{i}/project",
                "grade": f"{70+i}/100",
                "summary": f"Score: {70+i}/100. Test summary.",
                "notes": ""
            }
            for i in range(30)
        ]

        filepath = populate_excel(student_data, "Large Test", temp_dir)

        wb = load_workbook(filepath)
        ws = wb.active

        assert ws.max_row == 31  # Header + 30 students

    def test_header_formatting(self, temp_dir, sample_student_data):
        """Test header row formatting (bold, color, alignment)."""
        filepath = populate_excel(sample_student_data, "Test", temp_dir)

        wb = load_workbook(filepath)
        ws = wb.active

        for col_idx in range(1, 9):
            cell = ws.cell(row=1, column=col_idx)

            assert cell.font.bold is True
            assert cell.font.size == 11
            assert cell.font.name == 'Calibri'
            assert cell.fill.start_color.rgb == "FFD9E1F2" or cell.fill.start_color.rgb == "D9E1F2"
            assert cell.alignment.horizontal == "center"
            assert cell.border.left.style == 'thin'
            assert cell.border.top.style == 'thin'

    def test_column_widths(self, temp_dir, sample_student_data):
        """Test column widths are set correctly."""
        filepath = populate_excel(sample_student_data, "Test", temp_dir)

        wb = load_workbook(filepath)
        ws = wb.active

        assert ws.column_dimensions['A'].width == 12  # Student ID
        assert ws.column_dimensions['B'].width == 20  # Student Name
        assert ws.column_dimensions['G'].width == 60  # Summary
        assert ws.column_dimensions['H'].width == 15  # Notes

    def test_hyperlink_creation(self, temp_dir, sample_student_data):
        """Test GitHub URL is clickable hyperlink."""
        filepath = populate_excel(sample_student_data, "Test", temp_dir)

        wb = load_workbook(filepath)
        ws = wb.active

        github_cell = ws['E2']  # GitHub URL column, first student
        assert github_cell.value == "Link"
        assert github_cell.hyperlink is not None
        assert "github.com" in github_cell.hyperlink.target
        assert github_cell.font.color.rgb == "FF0563C1" or github_cell.font.color.rgb == "0563C1"
        assert github_cell.font.underline == "single"

    def test_text_wrapping(self, temp_dir, sample_student_data):
        """Test Summary column has text wrapping enabled."""
        filepath = populate_excel(sample_student_data, "Test", temp_dir)

        wb = load_workbook(filepath)
        ws = wb.active

        summary_cell = ws['G2']  # Summary column, first student
        assert summary_cell.alignment.wrap_text is True

    def test_borders(self, temp_dir, sample_student_data):
        """Test all cells have borders."""
        filepath = populate_excel(sample_student_data, "Test", temp_dir)

        wb = load_workbook(filepath)
        ws = wb.active

        for row in range(1, 4):
            for col in range(1, 9):
                cell = ws.cell(row=row, column=col)
                assert cell.border.left.style == 'thin'
                assert cell.border.right.style == 'thin'
                assert cell.border.top.style == 'thin'
                assert cell.border.bottom.style == 'thin'

    def test_empty_student_data(self, temp_dir):
        """Test error raised for empty student data."""
        with pytest.raises(ValueError, match="student_data cannot be empty"):
            populate_excel([], "Test", temp_dir)

    def test_none_student_data(self, temp_dir):
        """Test error raised for None student data."""
        with pytest.raises(ValueError, match="student_data cannot be empty"):
            populate_excel(None, "Test", temp_dir)

    def test_output_directory_creation(self, temp_dir):
        """Test output directory created if missing."""
        nested_dir = os.path.join(temp_dir, "nested", "output")
        student_data = [
            {
                "student_id": "12345678",
                "student_name": "John Doe",
                "partner_name": "Jane Smith",
                "assignment_name": "Test",
                "github_url": "https://github.com/test",
                "grade": "85/100",
                "summary": "Score: 85/100. Good.",
                "notes": ""
            }
        ]

        filepath = populate_excel(student_data, "Test", nested_dir)

        assert os.path.exists(nested_dir)
        assert os.path.exists(filepath)

    def test_filename_sanitization(self, temp_dir, sample_student_data):
        """Test special characters removed from filename."""
        filepath = populate_excel(
            sample_student_data,
            "Design Patterns - Fall 2025 (CS401)",
            temp_dir
        )

        assert "FinalFeedback_DesignPatterns-Fall2025CS401.xlsx" in filepath
        assert "(" not in filepath
        assert ")" not in filepath

    def test_missing_github_url(self, temp_dir):
        """Test handling of missing GitHub URL."""
        student_data = [
            {
                "student_id": "12345678",
                "student_name": "John Doe",
                "partner_name": "Jane Smith",
                "assignment_name": "Test",
                "github_url": "",
                "grade": "0/100",
                "summary": "Error: GitHub URL not found.",
                "notes": "PDF_MISSING"
            }
        ]

        filepath = populate_excel(student_data, "Test", temp_dir)

        wb = load_workbook(filepath)
        ws = wb.active

        github_cell = ws['E2']
        assert github_cell.value == ""
        assert github_cell.hyperlink is None

    def test_notes_column_populated(self, temp_dir):
        """Test Notes column contains manual review flags."""
        student_data = [
            {
                "student_id": "12345678",
                "student_name": "John Doe",
                "partner_name": "NOT_FOUND",
                "assignment_name": "Test",
                "github_url": "https://github.com/test",
                "grade": "75/100",
                "summary": "Score: 75/100.",
                "notes": "LOW_CONFIDENCE_REVIEW"
            }
        ]

        filepath = populate_excel(student_data, "Test", temp_dir)

        wb = load_workbook(filepath)
        ws = wb.active

        notes_cell = ws['H2']
        assert notes_cell.value == "LOW_CONFIDENCE_REVIEW"
