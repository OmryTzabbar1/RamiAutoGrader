"""
Integration tests for excel-feedback-generator agent.

Tests the complete workflow:
1. Extract metadata from PDFs
2. Generate summaries from grading results
3. Create formatted Excel file

These tests use sample student data in tests/fixtures/
"""

import os
import pytest
import json
import tempfile
import shutil
from pathlib import Path
import sys

# Add skills to path
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '../../skills/extract-pdf-metadata/scripts'))
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '../../skills/generate-summary/scripts'))
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '../../skills/populate-excel/scripts'))

from extract_metadata import extract_metadata
from summarize_report import generate_summary
from create_excel import populate_excel


class TestFullWorkflow:
    """Test complete ExcelFeedback workflow with sample data."""

    @pytest.fixture
    def fixtures_dir(self):
        """Path to test fixtures directory."""
        return Path(__file__).parent.parent / "fixtures" / "sample_students"

    @pytest.fixture
    def temp_output_dir(self):
        """Create temporary directory for test outputs."""
        temp_dir = tempfile.mkdtemp()
        yield temp_dir
        shutil.rmtree(temp_dir)

    def test_full_workflow_three_students(self, fixtures_dir, temp_output_dir):
        """
        Test complete workflow with 3 students.

        Workflow:
        1. Extract metadata from each PDF
        2. Generate summary for each grading report
        3. Create Excel file with all data
        """
        students = ["john_doe", "alice_johnson", "charlie_brown"]
        student_data = []

        # Step 1 & 2: Process each student
        for student_name in students:
            student_dir = fixtures_dir / student_name

            # Load existing metadata (GitHub URL)
            metadata_path = student_dir / "metadata.json"
            with open(metadata_path, 'r') as f:
                metadata = json.load(f)

            # Load grading results
            grading_path = student_dir / "grading_results.json"
            with open(grading_path, 'r') as f:
                grading_results = json.load(f)

            # For this test, we'll create mock PDF metadata since we don't have real PDFs
            # In real usage, this would come from extract-pdf-metadata skill
            if student_name == "john_doe":
                pdf_metadata = {
                    "student_id": "12345678",
                    "student_name": "John Doe",
                    "partner_name": "Jane Smith",
                    "assignment_name": "Design Patterns",
                    "confidence": 0.95,
                    "extraction_status": "SUCCESS"
                }
            elif student_name == "alice_johnson":
                pdf_metadata = {
                    "student_id": "87654321",
                    "student_name": "Alice Johnson",
                    "partner_name": "Bob Williams",
                    "assignment_name": "Design Patterns",
                    "confidence": 1.0,
                    "extraction_status": "SUCCESS"
                }
            else:  # charlie_brown
                pdf_metadata = {
                    "student_id": "11223344",
                    "student_name": "Charlie Brown",
                    "partner_name": "NOT_FOUND",
                    "assignment_name": "Design Patterns",
                    "confidence": 0.6,
                    "extraction_status": "NEEDS_MANUAL_REVIEW"
                }

            # Generate summary (this will use mocked Claude API in unit tests,
            # but here we'll create a realistic summary)
            total = grading_results["total_score"]
            max_score = grading_results["max_score"]

            if student_name == "john_doe":
                summary = "Score: 77/100. Strong documentation (25/25) and excellent security (10/10). Needs improvement: 2 files exceed 150 lines, test coverage at 67%."
            elif student_name == "alice_johnson":
                summary = "Score: 92/100. Exceptional code quality (28/30) and perfect test coverage (15/15). Minor: could add 2 more ADRs to documentation."
            else:
                summary = "Score: 65/100. Good security practices (10/10) and naming conventions. Critical issues: missing PLANNING.md, only 8 commits, test coverage at 55%."

            # Combine all data
            student_entry = {
                "student_id": pdf_metadata["student_id"],
                "student_name": pdf_metadata["student_name"],
                "partner_name": pdf_metadata["partner_name"],
                "assignment_name": pdf_metadata["assignment_name"],
                "github_url": metadata["github_url"],
                "grade": f"{total}/{max_score}",
                "summary": summary,
                "notes": pdf_metadata["extraction_status"] if pdf_metadata["extraction_status"] != "SUCCESS" else ""
            }

            student_data.append(student_entry)

        # Step 3: Create Excel file
        excel_path = populate_excel(
            student_data=student_data,
            assignment_name="Design Patterns",
            output_dir=temp_output_dir
        )

        # Verify Excel file created
        assert os.path.exists(excel_path)
        assert "FinalFeedback_DesignPatterns.xlsx" in excel_path

        # Verify file size > 0
        assert os.path.getsize(excel_path) > 0

        # Load and verify Excel content
        from openpyxl import load_workbook
        wb = load_workbook(excel_path)
        ws = wb.active

        # Verify header row
        assert ws['A1'].value == "Student ID"
        assert ws['B1'].value == "Student Name"
        assert ws['G1'].value == "Summary"
        assert ws['H1'].value == "Notes"

        # Verify 3 students present (header + 3 rows)
        assert ws.max_row == 4

        # Verify student data
        assert ws['A2'].value == "12345678"  # John Doe
        assert ws['B2'].value == "John Doe"
        assert ws['A3'].value == "87654321"  # Alice Johnson
        assert ws['B3'].value == "Alice Johnson"
        assert ws['A4'].value == "11223344"  # Charlie Brown
        assert ws['B4'].value == "Charlie Brown"

        # Verify notes column (Charlie should have manual review flag)
        assert ws['H2'].value == ""  # John - no flag
        assert ws['H3'].value == ""  # Alice - no flag
        assert ws['H4'].value == "NEEDS_MANUAL_REVIEW"  # Charlie - flagged

        # Verify GitHub hyperlinks
        assert ws['E2'].hyperlink is not None
        assert "github.com/johndoe" in ws['E2'].hyperlink.target

    def test_workflow_with_missing_pdf(self, temp_output_dir):
        """
        Test workflow when one student's PDF is missing.

        Expected: Should create Excel with "PDF_MISSING" flag in Notes.
        """
        student_data = [
            {
                "student_id": "NOT_FOUND",
                "student_name": "NOT_FOUND",
                "partner_name": "NOT_FOUND",
                "assignment_name": "NOT_FOUND",
                "github_url": "",
                "grade": "0/100",
                "summary": "Error: PDF not found. Cannot generate feedback.",
                "notes": "PDF_MISSING"
            }
        ]

        excel_path = populate_excel(
            student_data=student_data,
            assignment_name="Test",
            output_dir=temp_output_dir
        )

        from openpyxl import load_workbook
        wb = load_workbook(excel_path)
        ws = wb.active

        assert ws['H2'].value == "PDF_MISSING"
        assert ws['A2'].value == "NOT_FOUND"

    def test_workflow_with_low_confidence_extraction(self, temp_output_dir):
        """
        Test workflow when PDF extraction has low confidence.

        Expected: Should create Excel with "LOW_CONFIDENCE_REVIEW" flag.
        """
        student_data = [
            {
                "student_id": "ABC123",  # Invalid format
                "student_name": "John",  # Single word
                "partner_name": "NOT_FOUND",
                "assignment_name": "DP",  # Too short
                "github_url": "https://github.com/test",
                "grade": "70/100",
                "summary": "Score: 70/100. Average performance across all categories.",
                "notes": "LOW_CONFIDENCE_REVIEW"
            }
        ]

        excel_path = populate_excel(
            student_data=student_data,
            assignment_name="Test",
            output_dir=temp_output_dir
        )

        from openpyxl import load_workbook
        wb = load_workbook(excel_path)
        ws = wb.active

        assert ws['H2'].value == "LOW_CONFIDENCE_REVIEW"
        assert ws['A2'].value == "ABC123"  # Invalid student ID present but flagged

    def test_workflow_performance_many_students(self, temp_output_dir):
        """
        Test workflow performance with 30 students.

        Expected: Should complete in reasonable time (< 5 seconds for Excel creation).
        """
        import time

        # Create 30 mock students
        student_data = []
        for i in range(30):
            student_data.append({
                "student_id": f"1234567{i}",
                "student_name": f"Student {i}",
                "partner_name": f"Partner {i}",
                "assignment_name": "Design Patterns",
                "github_url": f"https://github.com/student{i}/project",
                "grade": f"{70 + i}/100",
                "summary": f"Score: {70 + i}/100. Good work overall with room for improvement in testing.",
                "notes": ""
            })

        start_time = time.time()
        excel_path = populate_excel(
            student_data=student_data,
            assignment_name="Performance Test",
            output_dir=temp_output_dir
        )
        elapsed_time = time.time() - start_time

        # Verify Excel created
        assert os.path.exists(excel_path)

        # Verify performance (should be < 5 seconds)
        assert elapsed_time < 5.0

        # Verify all students present
        from openpyxl import load_workbook
        wb = load_workbook(excel_path)
        ws = wb.active

        assert ws.max_row == 31  # Header + 30 students

    def test_workflow_excel_formatting(self, temp_output_dir):
        """
        Test that Excel file has proper formatting.

        Expected:
        - Header row: Bold, blue background
        - Borders on all cells
        - GitHub URL is hyperlink
        - Summary column has text wrapping
        """
        student_data = [
            {
                "student_id": "12345678",
                "student_name": "Test Student",
                "partner_name": "Test Partner",
                "assignment_name": "Test Assignment",
                "github_url": "https://github.com/test/repo",
                "grade": "85/100",
                "summary": "Score: 85/100. This is a long summary that should wrap to multiple lines when displayed in Excel. It tests the text wrapping functionality.",
                "notes": ""
            }
        ]

        excel_path = populate_excel(
            student_data=student_data,
            assignment_name="Formatting Test",
            output_dir=temp_output_dir
        )

        from openpyxl import load_workbook
        wb = load_workbook(excel_path)
        ws = wb.active

        # Test header formatting
        header_cell = ws['A1']
        assert header_cell.font.bold is True
        assert header_cell.fill.start_color.rgb in ["FFD9E1F2", "D9E1F2"]
        assert header_cell.alignment.horizontal == "center"
        assert header_cell.border.left.style == 'thin'

        # Test hyperlink
        github_cell = ws['E2']
        assert github_cell.value == "Link"
        assert github_cell.hyperlink is not None
        assert github_cell.font.underline == "single"

        # Test summary text wrapping
        summary_cell = ws['G2']
        assert summary_cell.alignment.wrap_text is True

        # Test borders on data cells
        data_cell = ws['B2']
        assert data_cell.border.left.style == 'thin'
        assert data_cell.border.top.style == 'thin'
