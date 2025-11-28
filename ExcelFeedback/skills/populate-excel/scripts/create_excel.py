"""
Create formatted Excel workbooks with student feedback data.

This module generates professionally formatted .xlsx files with headers, borders,
column widths, text wrapping, and clickable hyperlinks for instructor review.
"""

import os
from typing import List, Dict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter


COLUMNS = [
    {"header": "Student ID", "width": 12, "key": "student_id"},
    {"header": "Student Name", "width": 20, "key": "student_name"},
    {"header": "Partner Name", "width": 20, "key": "partner_name"},
    {"header": "Assignment Name", "width": 25, "key": "assignment_name"},
    {"header": "GitHub URL", "width": 40, "key": "github_url"},
    {"header": "Grade", "width": 10, "key": "grade"},
    {"header": "Summary", "width": 60, "key": "summary"},
    {"header": "Notes", "width": 15, "key": "notes"}
]


def add_header_row(ws, columns: List[Dict]):
    """
    Add formatted header row to worksheet.

    Args:
        ws: openpyxl Worksheet object
        columns: List of column definitions
    """
    header_font = Font(bold=True, size=11, name='Calibri')
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for idx, col in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=idx)
        cell.value = col["header"]
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(idx)].width = col["width"]


def add_data_rows(ws, student_data: List[Dict], columns: List[Dict]):
    """
    Add student data rows to worksheet.

    Args:
        ws: openpyxl Worksheet object
        student_data: List of student dictionaries
        columns: List of column definitions
    """
    cell_font = Font(size=11, name='Calibri')
    cell_alignment = Alignment(horizontal="left", vertical="top", wrap_text=False)
    summary_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    hyperlink_font = Font(color="0563C1", underline="single", size=11)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for row_idx, student in enumerate(student_data, start=2):
        for col_idx, col in enumerate(columns, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            value = student.get(col["key"], "")

            if col["key"] == "github_url" and value:
                cell.hyperlink = value
                cell.value = "Link"
                cell.font = hyperlink_font
            else:
                cell.value = value
                cell.font = cell_font

            if col["key"] == "summary":
                cell.alignment = summary_alignment
            else:
                cell.alignment = cell_alignment

            cell.border = thin_border


def populate_excel(student_data: List[Dict], assignment_name: str, output_dir: str = "results/") -> str:
    """
    Create formatted Excel workbook with student feedback.

    Args:
        student_data: List of student dictionaries
        assignment_name: Name of assignment
        output_dir: Output directory path

    Returns:
        Path to created Excel file

    Raises:
        ValueError: If student_data is empty
        PermissionError: If cannot create output directory
        IOError: If cannot save Excel file
    """
    if not student_data or len(student_data) == 0:
        raise ValueError("student_data cannot be empty. Need at least 1 student.")

    try:
        os.makedirs(output_dir, exist_ok=True)
    except PermissionError:
        raise PermissionError(f"Cannot create output directory: {output_dir}. Check permissions.")

    wb = Workbook()
    ws = wb.active
    ws.title = "FinalFeedback"

    add_header_row(ws, COLUMNS)
    add_data_rows(ws, student_data, COLUMNS)

    safe_name = "".join(c for c in assignment_name if c.isalnum() or c in (' ', '-', '_'))
    safe_name = safe_name.replace(' ', '')
    filename = f"FinalFeedback_{safe_name}.xlsx"
    filepath = os.path.join(output_dir, filename)

    try:
        wb.save(filepath)
    except Exception as e:
        raise IOError(f"Failed to save Excel file to {filepath}: {e}")

    return filepath
